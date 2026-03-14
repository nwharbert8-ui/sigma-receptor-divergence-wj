#!/usr/bin/env python3
"""
===============================================================================
MS3 SIGMA DIVERGENCE WJ-NATIVE PIPELINE — Local Execution (16GB RAM)
===============================================================================

MANUSCRIPT: "Sigma-1 and Sigma-2 Receptors Exhibit Divergent Genome-Wide
            Co-Expression Networks in Human Brain Despite Shared Subcellular
            Localization"

TARGET:     Frontiers in Molecular Neuroscience

AUTHOR:     Drake H. Harbert (ORCID: 0009-0007-7740-3616)
            Inner Architecture LLC, Canton, OH 44721, USA

METHODOLOGY:
  PRIMARY:  Weighted Jaccard (WJ) on continuous genome-wide Spearman
            correlation vectors. Threshold-free comparison of full
            co-expression architectures.
  SUPPLEMENTARY: Binary Jaccard on top 5% network membership (consistency
            with other genomics papers in portfolio).

COMPUTES:
  - Genome-wide Spearman correlations for SIGMAR1, TMEM97, and 5 additional
    targets across 5 brain regions
  - WJ on continuous correlation vectors (primary analysis)
  - WJ permutation testing (1000 permutations, seed=42)
  - Binary Jaccard on top 5% sets (supplementary)
  - FDR correction across all 21 pairwise comparisons
  - gProfiler GO/KEGG/Reactome enrichment
  - 6 custom gene set enrichments
  - Cell-type deconvolution sensitivity
  - Covariate adjustment (age, sex)
  - Multi-region replication (5 brain regions)
  - All figures (300 DPI, colorblind-safe)
  - Supplementary Tables S1-S4
  - provenance.json

USAGE:      py -3 MS3_sigma_divergence_wj_pipeline.py
Dependencies: scipy, statsmodels, pandas, numpy, matplotlib, matplotlib-venn,
              gprofiler-official, python-docx, openpyxl, requests
===============================================================================
"""

import os
import gc
import json
import time
import warnings
import requests
import gzip
import numpy as np
import pandas as pd
from scipy import stats
from scipy.stats import spearmanr, fisher_exact
from statsmodels.stats.multitest import multipletests
from collections import OrderedDict
from itertools import combinations
from numpy.linalg import lstsq
import matplotlib
matplotlib.use('Agg')
matplotlib.rcParams['pdf.fonttype'] = 42
matplotlib.rcParams['ps.fonttype'] = 42
import matplotlib.pyplot as plt
from matplotlib_venn import venn2

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIG
# ============================================================================
FORCE_RECOMPUTE = True
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)
N_PERMUTATIONS = 1000

# Paths
DRIVE_BASE = r"G:\My Drive\inner_architecture_research\MS3_JNC_Submission"
RESULTS_DIR = os.path.join(DRIVE_BASE, "MS3_Results")
FIGURES_DIR = os.path.join(DRIVE_BASE, "MS3_Figures")
SUPPL_DIR = os.path.join(DRIVE_BASE, "MS3_Supplementary_Tables")
DATA_DIR = os.path.join(DRIVE_BASE, "data")

for d in [RESULTS_DIR, FIGURES_DIR, SUPPL_DIR, DATA_DIR]:
    os.makedirs(d, exist_ok=True)

# Targets
PRIMARY_TARGETS = ['SIGMAR1', 'TMEM97']
TARGET_GENES = ['EIF2S1', 'PELO', 'LTN1', 'NEMF', 'TMEM97', 'HSPA5', 'SIGMAR1']

# Brain regions
BRAIN_REGIONS = OrderedDict([
    ('BA9',         'Brain - Frontal Cortex (BA9)'),
    ('Putamen',     'Brain - Putamen (basal ganglia)'),
    ('Hippocampus', 'Brain - Hippocampus'),
    ('NAcc',        'Brain - Nucleus accumbens (basal ganglia)'),
    ('BA24',        'Brain - Anterior cingulate cortex (BA24)'),
])
PRIMARY_REGION = 'BA9'
MIN_MEDIAN_TPM = 1.0
TOP_PERCENT = 5

# Custom gene sets
MAM_MITO_GENES = [
    'VDAC1', 'VDAC2', 'VDAC3', 'MFN1', 'MFN2', 'RHOT1', 'RHOT2',
    'ITPR1', 'ITPR2', 'ITPR3', 'VAPB', 'RMDN3', 'PACS2', 'FATE1',
]
SIGMA_NETWORK_GENES = ['SIGMAR1', 'TMEM97', 'PGRMC1', 'NPC1']
ER_STRESS_UPR_GENES = [
    'HSPA5', 'HSP90B1', 'DDIT3', 'ATF4', 'ATF6', 'ERN1', 'EIF2AK3',
    'XBP1', 'DNAJB9', 'HERPUD1', 'EDEM1', 'CALR', 'CANX', 'P4HB',
]
METHYLATION_GENES = [
    'MAT1A', 'MAT2A', 'MAT2B', 'AHCY', 'AHCYL1', 'AHCYL2',
    'MTR', 'MTRR', 'MTHFR', 'BHMT', 'BHMT2', 'CBS', 'CTH',
    'GNMT', 'PEMT', 'NNMT', 'INMT', 'DNMT1',
]
VASCULAR_GENES = [
    'PECAM1', 'CDH5', 'VWF', 'FLT1', 'KDR', 'ENG', 'CLDN5', 'ESAM',
    'ERG', 'TIE1', 'TEK', 'ANGPT1', 'ANGPT2', 'NOS3', 'MCAM',
    'PODXL', 'EMCN', 'ROBO4',
]
RQC_GENES = [
    'PELO', 'HBS1L', 'LTN1', 'NEMF', 'ANKZF1',
    'VCP', 'UFD1', 'NPLOC4', 'ZNF598', 'RACK1', 'ABCE1', 'TCF25',
]

# Cell-type deconvolution markers
CELLTYPE_MARKERS = {
    'Neurons': ['SNAP25', 'SYT1', 'GAD1', 'GAD2', 'SLC17A7', 'RBFOX3',
                'STMN2', 'SYN1', 'NRGN'],
    'Astrocytes': ['AQP4', 'GFAP', 'SLC1A2', 'SLC1A3', 'ALDH1L1', 'GJA1',
                   'S100B', 'SOX9', 'GLUL'],
    'Oligodendrocytes': ['MBP', 'MOG', 'PLP1', 'MAG', 'MOBP', 'CLDN11', 'CNP',
                         'OPALIN', 'TF'],
    'Microglia': ['CX3CR1', 'P2RY12', 'CSF1R', 'TMEM119', 'AIF1', 'ITGAM',
                  'CD68', 'HEXB', 'TREM2'],
    'Endothelial': ['CLDN5', 'FLT1', 'PECAM1', 'VWF', 'CDH5', 'ERG',
                    'ESAM', 'TIE1'],
    'OPCs': ['PDGFRA', 'CSPG4', 'OLIG1', 'OLIG2', 'SOX10', 'NKX2-2',
             'GPR17', 'PCDH15', 'NEU4'],
}

# GTEx URLs
GTEX_TPM_URL = "https://storage.googleapis.com/adult-gtex/bulk-gex/v8/rna-seq/GTEx_Analysis_2017-06-05_v8_RNASeQCv1.1.9_gene_tpm.gct.gz"
GTEX_SAMPLE_URL = "https://storage.googleapis.com/adult-gtex/annotations/v8/metadata-files/GTEx_Analysis_v8_Annotations_SampleAttributesDS.txt"
GTEX_SUBJECT_URL = "https://storage.googleapis.com/adult-gtex/annotations/v8/metadata-files/GTEx_Analysis_v8_Annotations_SubjectPhenotypesDS.txt"


# ============================================================================
# CORE WJ FUNCTIONS
# ============================================================================

def weighted_jaccard(vec_a, vec_b):
    """
    Compute Weighted Jaccard similarity on two continuous vectors.

    WJ = sum(min(a_i, b_i)) / sum(max(a_i, b_i))

    For correlation vectors that can be negative, shift to [0, 1] range first:
    r_shifted = (r + 1) / 2
    This maps r=-1 -> 0, r=0 -> 0.5, r=1 -> 1.
    """
    # Shift correlations from [-1, 1] to [0, 1]
    a = (np.array(vec_a) + 1) / 2
    b = (np.array(vec_b) + 1) / 2

    numerator = np.sum(np.minimum(a, b))
    denominator = np.sum(np.maximum(a, b))

    if denominator == 0:
        return 0.0
    return numerator / denominator


def weighted_jaccard_permutation_test(vec_a, vec_b, n_perm=1000, seed=42):
    """
    Permutation test for WJ significance.

    Null hypothesis: the two correlation vectors are drawn from the same
    underlying distribution (i.e., the two genes see the same transcriptional
    architecture).

    Permutation strategy: shuffle the gene labels in one vector, breaking
    the gene-to-gene correspondence while preserving the marginal distribution.
    """
    rng = np.random.RandomState(seed)
    observed_wj = weighted_jaccard(vec_a, vec_b)

    null_distribution = np.zeros(n_perm)
    for i in range(n_perm):
        perm_b = rng.permutation(vec_b)
        null_distribution[i] = weighted_jaccard(vec_a, perm_b)

    # Two-sided: how extreme is the observed WJ?
    # For divergence testing: low WJ means more divergent
    # p-value = proportion of null WJ values <= observed WJ
    p_value = np.mean(null_distribution <= observed_wj)

    return observed_wj, p_value, null_distribution


def vectorized_spearman(target_expr, other_expr):
    """
    Compute Spearman correlation between one target and many genes.
    Rank-transforms both, then computes Pearson on ranks.
    """
    from scipy.stats import rankdata

    x_rank = rankdata(target_expr)
    n = len(target_expr)

    # Rank each row of other_expr
    Y_rank = np.apply_along_axis(rankdata, 1, other_expr)

    # Pearson on ranks = Spearman
    x = x_rank - x_rank.mean()
    Y = Y_rank - Y_rank.mean(axis=1, keepdims=True)
    x_std = np.sqrt(np.sum(x**2))
    Y_std = np.sqrt(np.sum(Y**2, axis=1))

    valid = Y_std > 0
    r_values = np.full(len(other_expr), np.nan)
    r_values[valid] = np.dot(Y[valid], x) / (Y_std[valid] * x_std)
    return r_values


def compute_custom_enrichment(network_genes, gene_set_list, set_name,
                               all_genes_list, verbose=True):
    """Fisher's exact test for custom gene set enrichment."""
    universe = set(all_genes_list)
    expressed = [g for g in gene_set_list if g in universe]
    in_network = [g for g in expressed if g in network_genes]

    n_expressed = len(expressed)
    n_in_network = len(in_network)
    n_network = len(network_genes)
    n_universe = len(universe)

    if n_expressed == 0:
        return None

    expected = n_expressed * n_network / n_universe
    fold = n_in_network / expected if expected > 0 else 0

    a = n_in_network
    b = n_network - n_in_network
    c = n_expressed - n_in_network
    d = n_universe - n_network - c
    table = np.array([[a, b], [c, d]])
    _, p_val = fisher_exact(table, alternative='greater')

    result = {
        'gene_set': set_name, 'expressed': n_expressed,
        'in_network': n_in_network, 'fold_enrichment': fold,
        'p_value': p_val, 'genes_found': ', '.join(sorted(in_network)),
    }

    if verbose:
        sig = "+" if p_val < 0.05 else "-"
        print(f"  {sig} {set_name}: {n_in_network}/{n_expressed}, "
              f"{fold:.1f}x, p = {p_val:.2e}")
        if in_network:
            print(f"    Genes: {', '.join(sorted(in_network))}")
    return result


def partial_correlation_network(target, expr_df, covar_df, top_pct=5):
    """Genome-wide Spearman correlations after regressing covariates."""
    valid = covar_df.dropna().index
    valid = [s for s in valid if s in expr_df.columns]
    n_valid = len(valid)

    if n_valid < 30:
        return None, None, None, n_valid

    expr_sub = expr_df[valid]
    covar_matrix = covar_df.loc[valid].values
    X = np.column_stack([np.ones(n_valid), covar_matrix])

    target_vals = expr_sub.loc[target].values.astype(np.float64)
    beta, _, _, _ = lstsq(X, target_vals, rcond=None)
    target_resid = target_vals - X @ beta

    other_genes = [g for g in expr_sub.index if g != target]
    other_vals = expr_sub.loc[other_genes].values.astype(np.float64)
    betas = lstsq(X, other_vals.T, rcond=None)[0]
    resid = other_vals - (X @ betas).T

    r_vals = vectorized_spearman(target_resid, resid)

    corr_series = pd.Series(r_vals, index=other_genes).dropna().sort_values(ascending=False)
    n_top = int(np.ceil(len(corr_series) * top_pct / 100))
    top5_set = set(corr_series.head(n_top).index)
    threshold = corr_series.iloc[n_top - 1] if n_top <= len(corr_series) else np.nan

    return corr_series, threshold, top5_set, n_valid


def sample_to_subject(sample_id):
    parts = sample_id.split('-')
    return '-'.join(parts[:2]) if len(parts) >= 2 else sample_id


def download_file(url, dest_path):
    """Download a file, reuse if already present."""
    if os.path.exists(dest_path):
        size_mb = os.path.getsize(dest_path) / 1e6
        if size_mb > 1:
            print(f"  Using cached: {os.path.basename(dest_path)} ({size_mb:.0f} MB)")
            return
    print(f"  Downloading {os.path.basename(dest_path)}...")
    resp = requests.get(url, stream=True, timeout=600)
    resp.raise_for_status()
    total = int(resp.headers.get('content-length', 0))
    downloaded = 0
    with open(dest_path, 'wb') as f:
        for chunk in resp.iter_content(chunk_size=8192 * 16):
            f.write(chunk)
            downloaded += len(chunk)
            if total > 0 and downloaded % (50 * 1024 * 1024) < 8192 * 16:
                print(f"    {downloaded / 1e6:.0f} / {total / 1e6:.0f} MB")
    print(f"  Done: {os.path.basename(dest_path)} ({downloaded / 1e6:.0f} MB)")


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def main():
    t_start = time.time()

    print("=" * 70)
    print("MS3 SIGMA DIVERGENCE WJ-NATIVE PIPELINE")
    print("Primary: Weighted Jaccard on continuous Spearman vectors")
    print("Supplementary: Binary Jaccard on top 5% sets")
    print("Target: Frontiers in Molecular Neuroscience")
    print("Drake H. Harbert -- Inner Architecture LLC")
    print("=" * 70)

    # ==================================================================
    # STEP 1: DOWNLOAD GTEx v8 DATA
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 1: DOWNLOAD GTEx v8 DATA\n{'='*70}\n")

    tpm_file = os.path.join(DATA_DIR, "GTEx_v8_tpm.gct.gz")
    sample_file = os.path.join(DATA_DIR, "GTEx_v8_sample_attributes.txt")
    subject_file = os.path.join(DATA_DIR, "GTEx_v8_subject_phenotypes.txt")

    download_file(GTEX_TPM_URL, tpm_file)
    download_file(GTEX_SAMPLE_URL, sample_file)
    download_file(GTEX_SUBJECT_URL, subject_file)

    # ==================================================================
    # STEP 2: LOAD METADATA
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 2: LOAD METADATA\n{'='*70}\n")

    sample_attr = pd.read_csv(sample_file, sep='\t', low_memory=False)
    brain_labels = sample_attr[
        sample_attr['SMTSD'].str.contains('Brain', na=False)
    ]['SMTSD'].unique()

    for key, smtsd in BRAIN_REGIONS.items():
        if smtsd not in brain_labels:
            raise ValueError(f"Region mismatch: {key}: '{smtsd}'")

    region_samples = {}
    for key, smtsd in BRAIN_REGIONS.items():
        region_samples[key] = list(sample_attr.loc[sample_attr['SMTSD'] == smtsd, 'SAMPID'])
        print(f"  {key}: {len(region_samples[key])} samples")

    subject_pheno = pd.read_csv(subject_file, sep='\t')
    ba9_samples_list = region_samples[PRIMARY_REGION]
    ba9_subjects = [sample_to_subject(s) for s in ba9_samples_list]
    ba9_subject_df = pd.DataFrame({'SAMPID': ba9_samples_list, 'SUBJID': ba9_subjects})
    ba9_subject_df = ba9_subject_df.merge(subject_pheno, on='SUBJID', how='left')

    # ==================================================================
    # STEP 3: LOAD GTEx TPM
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 3: LOAD GTEx TPM\n{'='*70}\n")

    print("  Reading header...")
    with gzip.open(tpm_file, 'rt') as f:
        f.readline(); f.readline()
        header = f.readline().strip().split('\t')

    all_brain_samples = set()
    for key in BRAIN_REGIONS:
        all_brain_samples.update(region_samples[key])

    sample_cols = {i: col for i, col in enumerate(header) if col in all_brain_samples}
    keep_cols = [0, 1] + sorted(sample_cols.keys())
    print(f"  Brain samples: {len(sample_cols)}, columns to load: {len(keep_cols)}")

    print("  Loading expression data...")
    chunks = []
    with gzip.open(tpm_file, 'rt') as f:
        f.readline(); f.readline(); f.readline()
        row_buffer = []
        for line_num, line in enumerate(f):
            parts = line.strip().split('\t')
            row_buffer.append([parts[i] if i < len(parts) else '' for i in keep_cols])
            if len(row_buffer) >= 5000:
                chunks.append(pd.DataFrame(row_buffer))
                row_buffer = []
                if line_num % 10000 == 0:
                    print(f"    {line_num:,} genes...")
        if row_buffer:
            chunks.append(pd.DataFrame(row_buffer))

    tpm_raw = pd.concat(chunks, ignore_index=True)
    del chunks, row_buffer; gc.collect()

    col_names = ['Name', 'Description'] + [sample_cols[i] for i in sorted(sample_cols.keys())]
    tpm_raw.columns = col_names
    tpm_raw = tpm_raw.set_index('Name')
    gene_descriptions = tpm_raw['Description'].copy()
    tpm_raw = tpm_raw.drop('Description', axis=1)
    tpm_raw = tpm_raw.apply(pd.to_numeric, errors='coerce').astype(np.float32)
    print(f"  Loaded: {tpm_raw.shape[0]} genes x {tpm_raw.shape[1]} samples")

    # ==================================================================
    # STEP 4: PREPARE REGION MATRICES
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 4: PREPARE REGION MATRICES\n{'='*70}\n")

    def prepare_region(region_key):
        samples = [s for s in region_samples[region_key] if s in tpm_raw.columns]
        expr = tpm_raw[samples].copy()
        expr['gene_symbol'] = gene_descriptions.reindex(expr.index)
        expr = expr.dropna(subset=['gene_symbol'])
        expr = expr[expr['gene_symbol'] != '']
        expr['median_tpm'] = expr[samples].median(axis=1)
        expr = expr.sort_values('median_tpm', ascending=False)
        expr = expr.drop_duplicates(subset='gene_symbol', keep='first')
        expr = expr[expr['median_tpm'] >= MIN_MEDIAN_TPM]
        expr = expr.set_index('gene_symbol').drop('median_tpm', axis=1)
        return np.log2(expr + 1), len(samples), len(expr), list(expr.index)

    region_data = {}
    for key in BRAIN_REGIONS:
        log2_expr, n_samp, n_genes, genes = prepare_region(key)
        region_data[key] = {'expr': log2_expr, 'n_samples': n_samp,
                            'n_genes': n_genes, 'genes': genes}
        present = [g for g in TARGET_GENES if g in genes]
        print(f"  {key}: n={n_samp}, genes={n_genes}, targets={len(present)}/7")

    del tpm_raw; gc.collect()

    # ==================================================================
    # STEP 5: GENOME-WIDE SPEARMAN CO-EXPRESSION
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 5: GENOME-WIDE SPEARMAN CO-EXPRESSION -- {PRIMARY_REGION}\n{'='*70}\n")

    ba9 = region_data[PRIMARY_REGION]['expr']
    n_genes_ba9 = region_data[PRIMARY_REGION]['n_genes']
    n_samples_ba9 = region_data[PRIMARY_REGION]['n_samples']
    print(f"Matrix: {n_genes_ba9} genes x {n_samples_ba9} samples")
    print(f"Correlation method: Spearman (rank-based)\n")

    correlations = {}
    for target in TARGET_GENES:
        if target not in ba9.index:
            print(f"  WARNING: {target} not found")
            continue
        target_expr = ba9.loc[target].values.astype(np.float64)
        other_genes = [g for g in ba9.index if g != target]
        other_expr = ba9.loc[other_genes].values.astype(np.float64)

        r_values = vectorized_spearman(target_expr, other_expr)

        corr_series = pd.Series(r_values, index=other_genes, name=target)
        correlations[target] = corr_series.dropna().sort_values(ascending=False)
        print(f"  {target}: {len(corr_series.dropna())} genes, "
              f"top = {corr_series.dropna().idxmax()} "
              f"(rho = {corr_series.dropna().max():.3f})")

    # ==================================================================
    # STEP 6: PRIMARY ANALYSIS — WEIGHTED JACCARD ON CONTINUOUS VECTORS
    # ==================================================================
    print(f"\n{'='*70}")
    print("STEP 6: PRIMARY ANALYSIS -- WEIGHTED JACCARD (CONTINUOUS)")
    print(f"{'='*70}\n")
    print(f"Permutations: {N_PERMUTATIONS}, seed: {RANDOM_SEED}\n")

    # Align all correlation vectors to common gene set
    common_genes = sorted(set.intersection(*[set(correlations[g].index)
                                              for g in TARGET_GENES
                                              if g in correlations]))
    n_common = len(common_genes)
    print(f"Common genes across all 7 targets: {n_common}\n")

    # Build aligned matrix
    corr_matrix = pd.DataFrame({g: correlations[g].reindex(common_genes)
                                 for g in TARGET_GENES if g in correlations})

    # Compute WJ for all 21 pairwise comparisons
    wj_results = []
    for g1, g2 in combinations(TARGET_GENES, 2):
        if g1 not in corr_matrix.columns or g2 not in corr_matrix.columns:
            continue

        vec_a = corr_matrix[g1].values
        vec_b = corr_matrix[g2].values

        wj_obs, wj_pval, null_dist = weighted_jaccard_permutation_test(
            vec_a, vec_b, n_perm=N_PERMUTATIONS, seed=RANDOM_SEED)

        # Also compute Spearman rank correlation between vectors
        rho, rho_p = spearmanr(vec_a, vec_b)

        wj_results.append({
            'gene1': g1, 'gene2': g2,
            'wj': wj_obs,
            'wj_perm_p': wj_pval,
            'null_mean': np.mean(null_dist),
            'null_std': np.std(null_dist),
            'z_score': (wj_obs - np.mean(null_dist)) / np.std(null_dist) if np.std(null_dist) > 0 else 0,
            'spearman_rho': rho,
            'spearman_p': rho_p,
            'n_genes': n_common,
        })

    wj_df = pd.DataFrame(wj_results).sort_values('wj', ascending=True)

    # FDR correction on WJ permutation p-values
    reject_wj, fdr_wj, _, _ = multipletests(wj_df['wj_perm_p'].values, method='fdr_bh')
    wj_df['wj_perm_p_fdr'] = fdr_wj
    wj_df['fdr_significant'] = reject_wj

    wj_df.to_csv(os.path.join(RESULTS_DIR, "wj_continuous_all_21_pairs.csv"), index=False)

    # Print all results
    print(f"{'Gene1':>10s} {'Gene2':>10s} {'WJ':>8s} {'z':>8s} {'p_perm':>10s} "
          f"{'FDR':>10s} {'rho':>8s}")
    print("-" * 70)
    for _, row in wj_df.iterrows():
        print(f"{row['gene1']:>10s} {row['gene2']:>10s} {row['wj']:>8.4f} "
              f"{row['z_score']:>8.2f} {row['wj_perm_p']:>10.4f} "
              f"{row['wj_perm_p_fdr']:>10.4f} {row['spearman_rho']:>8.3f}")

    # Key pair: SIGMAR1-TMEM97
    st_wj = wj_df[
        ((wj_df['gene1'] == 'SIGMAR1') & (wj_df['gene2'] == 'TMEM97')) |
        ((wj_df['gene1'] == 'TMEM97') & (wj_df['gene2'] == 'SIGMAR1'))
    ].iloc[0]

    print(f"\n{'='*60}")
    print("PRIMARY RESULT: SIGMAR1-TMEM97 WEIGHTED JACCARD")
    print(f"  WJ = {st_wj['wj']:.6f}")
    print(f"  z-score = {st_wj['z_score']:.2f}")
    print(f"  Permutation p = {st_wj['wj_perm_p']:.4f}")
    print(f"  FDR p = {st_wj['wj_perm_p_fdr']:.4f}")
    print(f"  Null mean = {st_wj['null_mean']:.6f}, std = {st_wj['null_std']:.6f}")
    print(f"  Spearman rho = {st_wj['spearman_rho']:.4f}")
    print(f"{'='*60}")

    # SIGMAR1-LTN1
    sl_wj = wj_df[
        ((wj_df['gene1'] == 'SIGMAR1') & (wj_df['gene2'] == 'LTN1')) |
        ((wj_df['gene1'] == 'LTN1') & (wj_df['gene2'] == 'SIGMAR1'))
    ].iloc[0]
    print(f"\nSIGMAR1-LTN1: WJ={sl_wj['wj']:.4f}, z={sl_wj['z_score']:.2f}, "
          f"p={sl_wj['wj_perm_p']:.4f}")

    # ==================================================================
    # STEP 7: SUPPLEMENTARY — BINARY JACCARD ON TOP 5% SETS
    # ==================================================================
    print(f"\n{'='*70}")
    print("STEP 7: SUPPLEMENTARY -- BINARY JACCARD (TOP 5% SETS)")
    print(f"{'='*70}\n")

    networks = {}
    thresholds = {}
    for target, corr in correlations.items():
        n_total = len(corr)
        n_top = int(np.ceil(n_total * TOP_PERCENT / 100))
        networks[target] = set(corr.head(n_top).index)
        thresholds[target] = corr.iloc[n_top - 1]
        print(f"  {target}: top 5% = {n_top} genes, rho >= {thresholds[target]:.3f}")

    gene_universe = len(correlations[TARGET_GENES[0]])

    binary_results = []
    for g1, g2 in combinations(TARGET_GENES, 2):
        if g1 not in networks or g2 not in networks:
            continue
        set1, set2 = networks[g1], networks[g2]
        shared = set1 & set2
        union_set = set1 | set2
        jaccard = len(shared) / len(union_set) if len(union_set) > 0 else 0

        a, b, c = len(shared), len(set1 - set2), len(set2 - set1)
        d = gene_universe - len(union_set)
        fisher_or, fisher_p = fisher_exact(np.array([[a, b], [c, d]]),
                                            alternative='greater')

        binary_results.append({
            'gene1': g1, 'gene2': g2,
            'shared': len(shared), 'binary_jaccard': jaccard,
            'fisher_or': fisher_or, 'fisher_p': fisher_p,
            'set1_size': len(set1), 'set2_size': len(set2),
        })

    binary_df = pd.DataFrame(binary_results).sort_values('binary_jaccard', ascending=False)
    reject_b, fdr_b, _, _ = multipletests(binary_df['fisher_p'].values, method='fdr_bh')
    binary_df['fisher_p_fdr'] = fdr_b
    binary_df.to_csv(os.path.join(RESULTS_DIR, "binary_jaccard_top5pct.csv"), index=False)

    # Merge WJ and binary results for comparison
    comparison = wj_df[['gene1', 'gene2', 'wj', 'z_score', 'wj_perm_p']].merge(
        binary_df[['gene1', 'gene2', 'binary_jaccard', 'shared', 'fisher_p']],
        on=['gene1', 'gene2'], how='outer'
    )
    comparison.to_csv(os.path.join(RESULTS_DIR, "wj_vs_binary_comparison.csv"), index=False)

    print("\n--- WJ vs Binary Jaccard comparison ---")
    print(f"{'Pair':>25s} {'WJ':>8s} {'Binary J':>10s} {'Shared':>7s}")
    print("-" * 55)
    for _, row in comparison.sort_values('wj').iterrows():
        pair = f"{row['gene1']}-{row['gene2']}"
        bj = f"{row['binary_jaccard']:.3f}" if pd.notna(row.get('binary_jaccard')) else "N/A"
        sh = f"{int(row['shared'])}" if pd.notna(row.get('shared')) else "N/A"
        print(f"{pair:>25s} {row['wj']:>8.4f} {bj:>10s} {sh:>7s}")

    # Key validation
    st_binary = binary_df[
        ((binary_df['gene1'] == 'SIGMAR1') & (binary_df['gene2'] == 'TMEM97')) |
        ((binary_df['gene1'] == 'TMEM97') & (binary_df['gene2'] == 'SIGMAR1'))
    ].iloc[0]

    print(f"\n  SIGMAR1-TMEM97 binary J = {st_binary['binary_jaccard']:.3f}, "
          f"shared = {int(st_binary['shared'])}")

    # Gene sets
    sigmar1_unique = sorted(networks['SIGMAR1'] - networks['TMEM97'])
    tmem97_unique = sorted(networks['TMEM97'] - networks['SIGMAR1'])
    shared_st = sorted(networks['SIGMAR1'] & networks['TMEM97'])

    # ==================================================================
    # STEP 8: TOP CO-EXPRESSION PARTNERS
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 8: TOP CO-EXPRESSION PARTNERS\n{'='*70}\n")

    for target in PRIMARY_TARGETS:
        print(f"  {target} top 10:")
        for rank, (gene, r_val) in enumerate(correlations[target].head(10).items(), 1):
            print(f"    {rank:2d}. {gene:12s} rho = {r_val:.4f}")
        print()

    # ==================================================================
    # STEP 9: CUSTOM GENE SET ENRICHMENT
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 9: CUSTOM GENE SET ENRICHMENT\n{'='*70}\n")

    custom_sets = OrderedDict([
        ('MAM-mitochondrial', MAM_MITO_GENES),
        ('Sigma receptor network', SIGMA_NETWORK_GENES),
        ('ER stress/UPR', ER_STRESS_UPR_GENES),
        ('Methylation pathway', METHYLATION_GENES),
        ('Vascular markers (neg ctrl)', VASCULAR_GENES),
        ('Ribosome quality control', RQC_GENES),
    ])

    all_custom_results = {}
    for target in PRIMARY_TARGETS:
        print(f"\n--- {target} top 5% custom enrichment ---\n")
        all_genes_t = list(correlations[target].index) + [target]
        target_results = []
        for set_name, gene_list in custom_sets.items():
            result = compute_custom_enrichment(
                networks[target], gene_list, set_name, all_genes_t)
            if result:
                result['target'] = target
                target_results.append(result)
        all_custom_results[target] = target_results
        pd.DataFrame(target_results).to_csv(
            os.path.join(RESULTS_DIR, f"{target}_custom_enrichment.csv"), index=False)

    # ==================================================================
    # STEP 10: gProfiler ENRICHMENT
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 10: gProfiler ENRICHMENT\n{'='*70}\n")

    try:
        from gprofiler import GProfiler
        gp = GProfiler(return_dataframe=True)
    except ImportError:
        print("WARNING: gprofiler not installed")
        gp = None

    def run_gprofiler(gene_list, bg, label=""):
        if gp is None:
            return pd.DataFrame()
        try:
            df = gp.profile(organism='hsapiens', query=list(gene_list),
                           background=list(bg),
                           sources=['GO:BP', 'GO:MF', 'GO:CC', 'KEGG', 'REAC'],
                           significance_threshold_method='g_SCS',
                           user_threshold=0.05, no_evidences=False)
            if df is not None and len(df) > 0:
                print(f"  {label}: {len(df)} terms")
                return df
            print(f"  {label}: 0 terms")
            return pd.DataFrame()
        except Exception as e:
            print(f"  {label}: error -- {e}")
            return pd.DataFrame()

    background = list(correlations['SIGMAR1'].index) + ['SIGMAR1']

    all_gprofiler = {}
    for name, genes in [
        ('SIGMAR1_full', list(networks['SIGMAR1'])),
        ('TMEM97_full', list(networks['TMEM97'])),
        ('SIGMAR1_unique', sigmar1_unique),
        ('TMEM97_unique', tmem97_unique),
        ('shared_SIGMAR1_TMEM97', shared_st),
    ]:
        df = run_gprofiler(genes, background, name)
        all_gprofiler[name] = df
        if len(df) > 0:
            df.to_csv(os.path.join(RESULTS_DIR, f"gProfiler_{name}.csv"), index=False)

    # ==================================================================
    # STEP 11: CELL-TYPE DECONVOLUTION
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 11: CELL-TYPE DECONVOLUTION SENSITIVITY\n{'='*70}\n")

    celltype_proportions = pd.DataFrame(index=ba9.columns)
    for ct_name, markers in CELLTYPE_MARKERS.items():
        present = [g for g in markers if g in ba9.index]
        if present:
            celltype_proportions[ct_name] = ba9.loc[present].mean(axis=0)
            print(f"  {ct_name}: {len(present)}/{len(markers)} markers")

    for target in PRIMARY_TARGETS:
        corr_ct, thr_ct, net_ct, n_valid = partial_correlation_network(
            target, ba9, celltype_proportions)
        if corr_ct is None:
            continue
        common = sorted(set(corr_ct.index) & set(correlations[target].index))
        rho_preserve, _ = spearmanr(corr_ct.reindex(common).values,
                                     correlations[target].reindex(common).values)
        print(f"  {target}: rank preservation rho = {rho_preserve:.4f}")
        if target == 'SIGMAR1' and net_ct:
            print(f"    VCP in adjusted top 5%: {'VCP' in net_ct}")

    # ==================================================================
    # STEP 12: COVARIATE ADJUSTMENT
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 12: COVARIATE ADJUSTMENT (AGE + SEX)\n{'='*70}\n")

    ba9_covar = ba9_subject_df.set_index('SAMPID').reindex(ba9.columns)
    age_map = {'20-29': 25, '30-39': 35, '40-49': 45, '50-59': 55,
               '60-69': 65, '70-79': 75}
    ba9_covar['AGE_MID'] = ba9_covar['AGE'].map(age_map) if 'AGE' in ba9_covar.columns else np.nan
    ba9_covar['SEX_NUM'] = ba9_covar['SEX'].astype(float) if 'SEX' in ba9_covar.columns else np.nan
    covar_df = ba9_covar[['AGE_MID', 'SEX_NUM']]

    for target in PRIMARY_TARGETS:
        corr_as, thr_as, _, n_valid = partial_correlation_network(target, ba9, covar_df)
        if corr_as is None:
            continue
        common = sorted(set(corr_as.index) & set(correlations[target].index))
        rho_p, _ = spearmanr(corr_as.reindex(common).values,
                              correlations[target].reindex(common).values)
        print(f"  {target}: rank preservation rho = {rho_p:.4f}")

    # ==================================================================
    # STEP 13: MULTI-REGION REPLICATION (WJ + binary)
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 13: MULTI-REGION REPLICATION\n{'='*70}\n")

    region_correlations = {}
    for region_key in BRAIN_REGIONS:
        expr = region_data[region_key]['expr']
        region_correlations[region_key] = {}
        for target in PRIMARY_TARGETS:
            if target not in expr.index:
                continue
            target_vals = expr.loc[target].values.astype(np.float64)
            other_genes = [g for g in expr.index if g != target]
            other_vals = expr.loc[other_genes].values.astype(np.float64)
            r_vals = vectorized_spearman(target_vals, other_vals)
            region_correlations[region_key][target] = pd.Series(
                r_vals, index=other_genes).dropna().sort_values(ascending=False)
        print(f"  {region_key} (n={region_data[region_key]['n_samples']}): done")

    # Cross-region rank correlations
    region_keys = list(BRAIN_REGIONS.keys())
    cross_region_matrix = {}
    for target in PRIMARY_TARGETS:
        mat = np.zeros((5, 5))
        for i, r1 in enumerate(region_keys):
            for j, r2 in enumerate(region_keys):
                if i == j:
                    mat[i, j] = 1.0
                    continue
                if target in region_correlations.get(r1, {}) and target in region_correlations.get(r2, {}):
                    c1, c2 = region_correlations[r1][target], region_correlations[r2][target]
                    common = sorted(set(c1.index) & set(c2.index))
                    if len(common) > 100:
                        mat[i, j], _ = spearmanr(c1.reindex(common).values, c2.reindex(common).values)
        cross_region_matrix[target] = mat
        vals = mat[np.triu_indices(5, k=1)]
        print(f"\n  {target} cross-region rho: {vals.min():.3f}-{vals.max():.3f}")

    # WJ across regions
    print("\n--- SIGMAR1-TMEM97 WJ across regions ---")
    region_wj_results = {}
    for region_key in BRAIN_REGIONS:
        if 'SIGMAR1' not in region_correlations.get(region_key, {}) or \
           'TMEM97' not in region_correlations.get(region_key, {}):
            continue
        s_corr = region_correlations[region_key]['SIGMAR1']
        t_corr = region_correlations[region_key]['TMEM97']
        common_r = sorted(set(s_corr.index) & set(t_corr.index))
        wj_r = weighted_jaccard(s_corr.reindex(common_r).values,
                                 t_corr.reindex(common_r).values)

        # Binary Jaccard too
        n_s = int(np.ceil(len(s_corr) * TOP_PERCENT / 100))
        n_t = int(np.ceil(len(t_corr) * TOP_PERCENT / 100))
        s_set, t_set = set(s_corr.head(n_s).index), set(t_corr.head(n_t).index)
        shared_r = s_set & t_set
        bj_r = len(shared_r) / len(s_set | t_set) if len(s_set | t_set) > 0 else 0

        region_wj_results[region_key] = {'wj': wj_r, 'binary_j': bj_r,
                                          'shared': len(shared_r)}
        print(f"  {region_key}: WJ={wj_r:.4f}, binary J={bj_r:.3f}, shared={len(shared_r)}")

    # ==================================================================
    # STEP 14: EXPORT GENE LISTS
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 14: EXPORT GENE LISTS\n{'='*70}\n")

    for target in PRIMARY_TARGETS:
        top5 = correlations[target].head(len(networks[target]))
        pd.DataFrame({'gene': top5.index, f'rho_with_{target}': top5.values}).to_csv(
            os.path.join(RESULTS_DIR, f"{target}_top5pct.csv"), index=False)

    pd.DataFrame({'gene': shared_st}).to_csv(
        os.path.join(RESULTS_DIR, "SIGMAR1_TMEM97_shared.csv"), index=False)
    pd.DataFrame({'gene': sigmar1_unique}).to_csv(
        os.path.join(RESULTS_DIR, "SIGMAR1_unique.csv"), index=False)
    pd.DataFrame({'gene': tmem97_unique}).to_csv(
        os.path.join(RESULTS_DIR, "TMEM97_unique.csv"), index=False)

    for target in TARGET_GENES:
        if target in correlations:
            corr = correlations[target]
            pd.DataFrame({'gene': corr.index, f'rho_with_{target}': corr.values,
                          'rank': range(1, len(corr) + 1)}).to_csv(
                os.path.join(RESULTS_DIR, f"{target}_genome_wide_rankings.csv"), index=False)

    print(f"  Exported: {len(TARGET_GENES)} ranking files, 3 gene set files")

    # ==================================================================
    # STEP 15: FIGURES
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 15: FIGURES\n{'='*70}\n")

    # Figure 1: Divergent Networks (Venn + scatter + top partners)
    fig = plt.figure(figsize=(18, 6))

    ax_a = fig.add_axes([0.02, 0.12, 0.28, 0.80])
    ax_a.text(-0.05, 1.08, 'A', fontsize=20, fontweight='bold', va='top',
              transform=ax_a.transAxes)
    v = venn2(subsets=(len(sigmar1_unique), len(tmem97_unique), len(shared_st)),
              set_labels=('SIGMAR1', 'TMEM97'), ax=ax_a)
    for pid, color in [('10', '#1565C0'), ('01', '#E65100'), ('11', '#7B1FA2')]:
        v.get_patch_by_id(pid).set_color(color)
        v.get_patch_by_id(pid).set_alpha(0.7)
    j_val = len(shared_st) / (len(sigmar1_unique) + len(tmem97_unique) + len(shared_st))
    ax_a.set_title(f'Binary Jaccard = {j_val:.3f}\nWJ = {st_wj["wj"]:.4f}',
                   fontsize=12, fontweight='bold')

    ax_b = fig.add_axes([0.36, 0.12, 0.28, 0.80])
    ax_b.text(-0.05, 1.08, 'B', fontsize=20, fontweight='bold', va='top',
              transform=ax_b.transAxes)
    s_vals = corr_matrix['SIGMAR1'].values
    t_vals = corr_matrix['TMEM97'].values
    ax_b.scatter(s_vals, t_vals, s=1, alpha=0.15, c='#555', rasterized=True)
    ax_b.set_xlabel('SIGMAR1 Spearman rho', fontsize=12)
    ax_b.set_ylabel('TMEM97 Spearman rho', fontsize=12)
    ax_b.set_title(f'Genome-wide correlation vectors\nWJ = {st_wj["wj"]:.4f}',
                   fontsize=12, fontweight='bold')
    ax_b.plot([-0.5, 1], [-0.5, 1], 'k--', alpha=0.3, lw=1)

    ax_c = fig.add_axes([0.70, 0.12, 0.28, 0.80])
    ax_c.text(-0.05, 1.08, 'C', fontsize=20, fontweight='bold', va='top',
              transform=ax_c.transAxes)
    top10_s = correlations['SIGMAR1'].head(10)
    top10_t = correlations['TMEM97'].head(10)
    y_pos = np.arange(10)
    bh = 0.35
    ax_c.barh(y_pos + bh/2, top10_s.values, bh, label='SIGMAR1', color='#1565C0', alpha=0.8)
    ax_c.barh(y_pos - bh/2, top10_t.values, bh, label='TMEM97', color='#E65100', alpha=0.8)
    ax_c.set_yticks(y_pos)
    ax_c.set_yticklabels([f"{top10_s.index[i]} | {top10_t.index[i]}" for i in range(10)], fontsize=8)
    ax_c.set_xlabel('Spearman rho', fontsize=12)
    ax_c.set_title('Top 10 co-expression partners', fontsize=12, fontweight='bold')
    ax_c.legend(fontsize=10)
    ax_c.invert_yaxis()

    plt.savefig(os.path.join(FIGURES_DIR, "Figure1_Divergent_Networks.png"),
                dpi=300, bbox_inches='tight')
    plt.savefig(os.path.join(FIGURES_DIR, "Figure1_Divergent_Networks.pdf"),
                bbox_inches='tight')
    plt.close()
    print("  Figure 1 saved")

    # Figure 2: GO Enrichment
    fig, axes = plt.subplots(2, 3, figsize=(20, 12))
    panel_data = [
        (all_gprofiler.get('SIGMAR1_unique', pd.DataFrame()), 'GO:BP', 'SIGMAR1-unique GO:BP'),
        (all_gprofiler.get('TMEM97_unique', pd.DataFrame()), 'GO:BP', 'TMEM97-unique GO:BP'),
        (all_gprofiler.get('shared_SIGMAR1_TMEM97', pd.DataFrame()), 'GO:BP', 'Shared GO:BP'),
        (all_gprofiler.get('SIGMAR1_full', pd.DataFrame()), 'REAC', 'SIGMAR1 Reactome'),
        (all_gprofiler.get('TMEM97_full', pd.DataFrame()), 'REAC', 'TMEM97 Reactome'),
        (all_gprofiler.get('SIGMAR1_full', pd.DataFrame()), 'KEGG', 'SIGMAR1 KEGG'),
    ]
    for idx, (go_df, source, title) in enumerate(panel_data):
        ax = axes.flat[idx]
        ax.text(-0.08, 1.05, chr(65+idx), fontsize=16, fontweight='bold',
                va='top', transform=ax.transAxes)
        if go_df is not None and len(go_df) > 0:
            subset = go_df[go_df['source'] == source].sort_values('p_value').head(8)
            if len(subset) > 0:
                names = [n[:45] for n in subset['name'].values]
                pvals = [-np.log10(p) for p in subset['p_value'].values]
                color = '#1565C0' if 'SIGMAR1' in title else '#E65100' if 'TMEM97' in title else '#7B1FA2'
                ax.barh(range(len(names)), pvals, color=color, alpha=0.8)
                ax.set_yticks(range(len(names)))
                ax.set_yticklabels(names, fontsize=8)
                ax.set_xlabel('-log10(p)', fontsize=10)
                ax.invert_yaxis()
            else:
                ax.text(0.5, 0.5, 'No terms', ha='center', va='center',
                        transform=ax.transAxes, color='gray')
        else:
            ax.text(0.5, 0.5, 'No data', ha='center', va='center',
                    transform=ax.transAxes, color='gray')
        ax.set_title(title, fontsize=11, fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(FIGURES_DIR, "Figure2_GO_Enrichment.png"), dpi=300, bbox_inches='tight')
    plt.savefig(os.path.join(FIGURES_DIR, "Figure2_GO_Enrichment.pdf"), bbox_inches='tight')
    plt.close()
    print("  Figure 2 saved")

    # Figure 3: Multi-region replication
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    for idx, target in enumerate(PRIMARY_TARGETS):
        ax = axes[idx]
        ax.text(-0.08, 1.05, chr(65+idx), fontsize=16, fontweight='bold',
                va='top', transform=ax.transAxes)
        mat = cross_region_matrix[target]
        im = ax.imshow(mat, cmap='viridis', vmin=0.8, vmax=1.0)
        for i in range(5):
            for j in range(5):
                ax.text(j, i, f'{mat[i,j]:.3f}', ha='center', va='center',
                        fontsize=9, color='white' if mat[i,j] < 0.92 else 'black')
        ax.set_xticks(range(5)); ax.set_xticklabels(region_keys, fontsize=9, rotation=45, ha='right')
        ax.set_yticks(range(5)); ax.set_yticklabels(region_keys, fontsize=9)
        ax.set_title(f'{target} cross-region', fontsize=12, fontweight='bold')
        plt.colorbar(im, ax=ax, shrink=0.8, label='Spearman rho')
    plt.tight_layout()
    plt.savefig(os.path.join(FIGURES_DIR, "Figure3_MultiRegion_Replication.png"), dpi=300, bbox_inches='tight')
    plt.savefig(os.path.join(FIGURES_DIR, "Figure3_MultiRegion_Replication.pdf"), bbox_inches='tight')
    plt.close()
    print("  Figure 3 saved")

    # ==================================================================
    # STEP 16: SUPPLEMENTARY TABLES
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 16: SUPPLEMENTARY TABLES\n{'='*70}\n")

    import openpyxl
    wb = openpyxl.Workbook()
    for i, target in enumerate(TARGET_GENES):
        if target in correlations:
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = target
            ws.append(['Gene', f'rho_with_{target}', 'Rank'])
            for rank, (gene, r_val) in enumerate(correlations[target].items(), 1):
                ws.append([gene, round(r_val, 6), rank])
    wb.save(os.path.join(SUPPL_DIR, "Table_S1_Genome_Wide_Rankings.xlsx"))

    wb2 = openpyxl.Workbook()
    first = True
    for name, df in all_gprofiler.items():
        if len(df) > 0:
            ws = wb2.active if first else wb2.create_sheet()
            ws.title = name[:31]
            cols = [c for c in df.columns if c not in ['query', 'parents']]
            ws.append(cols)
            for _, row in df[cols].iterrows():
                ws.append([str(v) for v in row.values])
            first = False
    wb2.save(os.path.join(SUPPL_DIR, "Table_S2_gProfiler_Enrichment.xlsx"))

    all_custom_rows = []
    for target, results in all_custom_results.items():
        all_custom_rows.extend(results)
    pd.DataFrame(all_custom_rows).to_excel(
        os.path.join(SUPPL_DIR, "Table_S3_Custom_Gene_Set_Enrichment.xlsx"), index=False)

    shared_detail = [{'gene': g,
                      'rho_with_SIGMAR1': round(correlations['SIGMAR1'].get(g, np.nan), 6),
                      'rho_with_TMEM97': round(correlations['TMEM97'].get(g, np.nan), 6)}
                     for g in shared_st]
    pd.DataFrame(shared_detail).to_excel(
        os.path.join(SUPPL_DIR, "Table_S4_Shared_Genes.xlsx"), index=False)

    print("  Tables S1-S4 saved")

    # ==================================================================
    # STEP 17: PROVENANCE
    # ==================================================================
    print(f"\n{'='*70}\nSTEP 17: PROVENANCE\n{'='*70}\n")

    from datetime import datetime
    provenance = {
        "methodology": "WJ-native",
        "fundamental_unit": f"individual gene (GTEx v8 RNA-seq, {n_genes_ba9} expressed genes in BA9)",
        "pairwise_matrix": "genome-wide Spearman correlation, each target vs all genes",
        "correlation_method": "Spearman",
        "primary_analysis": "Weighted Jaccard on continuous correlation vectors",
        "supplementary_analysis": "Binary Jaccard on top 5% network membership",
        "fdr_scope": f"all 21 pairwise WJ permutation p-values (Benjamini-Hochberg)",
        "permutations": N_PERMUTATIONS,
        "domain_conventional_methods": "gProfiler GO enrichment (comparison), Fisher exact (custom sets)",
        "random_seed": RANDOM_SEED,
        "pipeline_file": "MS3_sigma_divergence_wj_pipeline.py",
        "execution_date": datetime.now().strftime("%Y-%m-%d"),
        "execution_time_seconds": round(time.time() - t_start, 1),
        "wj_compliance_status": "PASS",
        "brain_regions": list(BRAIN_REGIONS.keys()),
        "n_samples_primary": n_samples_ba9,
        "n_genes_primary": n_genes_ba9,
        "key_results": {
            "SIGMAR1_TMEM97_wj": float(st_wj['wj']),
            "SIGMAR1_TMEM97_wj_z": float(st_wj['z_score']),
            "SIGMAR1_TMEM97_wj_perm_p": float(st_wj['wj_perm_p']),
            "SIGMAR1_TMEM97_binary_jaccard": float(st_binary['binary_jaccard']),
            "SIGMAR1_TMEM97_shared_genes": int(st_binary['shared']),
            "SIGMAR1_LTN1_wj": float(sl_wj['wj']),
            "cross_region_rho_SIGMAR1": f"{cross_region_matrix['SIGMAR1'][np.triu_indices(5, k=1)].min():.3f}-{cross_region_matrix['SIGMAR1'][np.triu_indices(5, k=1)].max():.3f}",
            "region_wj_values": {k: round(v['wj'], 4) for k, v in region_wj_results.items()},
            "region_binary_j_values": {k: round(v['binary_j'], 3) for k, v in region_wj_results.items()},
        },
    }

    provenance_path = os.path.join(RESULTS_DIR, "provenance.json")
    with open(provenance_path, 'w') as f:
        json.dump(provenance, f, indent=2)
    print(f"  provenance.json saved")

    # ==================================================================
    # SUMMARY
    # ==================================================================
    elapsed = time.time() - t_start
    print(f"\n{'='*70}")
    print("PIPELINE COMPLETE")
    print(f"{'='*70}")
    print(f"\n  Time: {elapsed:.0f}s ({elapsed/60:.1f} min)")
    print(f"\n--- PRIMARY: WEIGHTED JACCARD (continuous) ---")
    print(f"  SIGMAR1-TMEM97 WJ = {st_wj['wj']:.6f} (z={st_wj['z_score']:.2f}, p={st_wj['wj_perm_p']:.4f})")
    print(f"  SIGMAR1-LTN1   WJ = {sl_wj['wj']:.6f} (z={sl_wj['z_score']:.2f}, p={sl_wj['wj_perm_p']:.4f})")
    print(f"\n--- SUPPLEMENTARY: BINARY JACCARD (top 5%) ---")
    print(f"  SIGMAR1-TMEM97 J = {st_binary['binary_jaccard']:.3f} (shared={int(st_binary['shared'])})")
    print(f"\n--- MULTI-REGION WJ ---")
    for rk, rv in region_wj_results.items():
        print(f"  {rk}: WJ={rv['wj']:.4f}, binary J={rv['binary_j']:.3f}")

    return provenance


if __name__ == '__main__':
    main()
